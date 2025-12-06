"""
Enhanced Excel service for multi-sheet support and advanced features.
"""
import os
import logging
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for advanced Excel file operations with multi-sheet support."""
    
    def __init__(self):
        """Initialize Excel service."""
        pass
    
    def get_sheet_info(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract sheet names and metadata from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of sheet information dictionaries
        """
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_info = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Count rows and columns (approximate for large files)
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Check if sheet has data (not just headers)
                has_data = max_row > 1 if max_row else False
                
                sheet_info.append({
                    'name': sheet_name,
                    'row_count': max_row or 0,
                    'column_count': max_col or 0,
                    'has_data': has_data,
                    'is_visible': sheet.sheet_state == 'visible'
                })
            
            workbook.close()
            return sheet_info
            
        except Exception as e:
            logger.error(f"Failed to get sheet info from {file_path}: {str(e)}")
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    def read_excel_file(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: Optional[int] = None
    ) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Read Excel file with support for multi-sheet selection.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Optional sheet name to read
            sheet_index: Optional sheet index to read (0-based)
            
        Returns:
            Tuple of (DataFrame, metadata_dict)
        """
        try:
            # Determine which sheet to read
            if sheet_name is not None:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                selected_sheet = sheet_name
            elif sheet_index is not None:
                df = pd.read_excel(file_path, sheet_name=sheet_index)
                # Get actual sheet name
                workbook = load_workbook(file_path, read_only=True)
                selected_sheet = workbook.sheetnames[sheet_index]
                workbook.close()
            else:
                # Read first sheet by default
                df = pd.read_excel(file_path, sheet_name=0)
                workbook = load_workbook(file_path, read_only=True)
                selected_sheet = workbook.sheetnames[0]
                workbook.close()
            
            # Get all sheet info
            sheet_info = self.get_sheet_info(file_path)
            
            metadata = {
                'selected_sheet': selected_sheet,
                'total_sheets': len(sheet_info),
                'all_sheets': sheet_info,
                'file_type': 'xlsx'
            }
            
            return df, metadata
            
        except Exception as e:
            logger.error(f"Failed to read Excel file {file_path}: {str(e)}")
            raise ValueError(f"Failed to read Excel file: {str(e)}")
    
    def write_excel_file(
        self,
        data: pd.DataFrame,
        file_path: str,
        sheet_name: str = 'Sheet1',
        formatting: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Write Excel file with formatting preservation.
        
        Args:
            data: DataFrame to write
            file_path: Output file path
            sheet_name: Sheet name for the data
            formatting: Optional formatting configuration
            
        Returns:
            Path to generated file
        """
        try:
            formatting = formatting or {}
            
            # Create Excel writer with openpyxl engine
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=formatting.get('include_index', False),
                    startrow=formatting.get('start_row', 0),
                    startcol=formatting.get('start_col', 0)
                )
                
                # Access workbook and sheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting if specified
                if formatting.get('auto_filter', False):
                    worksheet.auto_filter.ref = worksheet.dimensions
                
                if formatting.get('freeze_panes'):
                    freeze_cell = formatting['freeze_panes']
                    worksheet.freeze_panes = freeze_cell
                
                # Auto-adjust column widths
                if formatting.get('auto_width', True):
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.info(f"Generated Excel file: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Failed to write Excel file {file_path}: {str(e)}")
            raise ValueError(f"Failed to write Excel file: {str(e)}")
    
    def extract_excel_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Extract comprehensive metadata from Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with Excel metadata
        """
        try:
            workbook = load_workbook(file_path, read_only=True)
            
            # Get workbook properties
            props = workbook.properties
            
            metadata = {
                'creator': props.creator,
                'last_modified_by': props.lastModifiedBy,
                'created': props.created.isoformat() if props.created else None,
                'modified': props.modified.isoformat() if props.modified else None,
                'title': props.title,
                'subject': props.subject,
                'description': props.description,
                'keywords': props.keywords,
                'category': props.category,
                'sheets': self.get_sheet_info(file_path),
                'has_formulas': self._check_for_formulas(workbook),
                'has_charts': self._check_for_charts(workbook),
                'has_pivot_tables': self._check_for_pivot_tables(workbook)
            }
            
            workbook.close()
            return metadata
            
        except Exception as e:
            logger.error(f"Failed to extract Excel metadata from {file_path}: {str(e)}")
            return {
                'sheets': self.get_sheet_info(file_path),
                'error': str(e)
            }
    
    def _check_for_formulas(self, workbook) -> bool:
        """Check if workbook contains formulas."""
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(max_row=100):  # Check first 100 rows
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            return True
            return False
        except:
            return False
    
    def _check_for_charts(self, workbook) -> bool:
        """Check if workbook contains charts."""
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if hasattr(sheet, '_charts') and len(sheet._charts) > 0:
                    return True
            return False
        except:
            return False
    
    def _check_for_pivot_tables(self, workbook) -> bool:
        """Check if workbook contains pivot tables."""
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if hasattr(sheet, '_pivots') and len(sheet._pivots) > 0:
                    return True
            return False
        except:
            return False
    
    def convert_formulas_to_values(
        self,
        file_path: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        Convert all formulas in an Excel file to their calculated values.
        
        Args:
            file_path: Path to input Excel file
            output_path: Optional output path (defaults to input path with _values suffix)
            
        Returns:
            Path to output file
        """
        try:
            if output_path is None:
                base, ext = os.path.splitext(file_path)
                output_path = f"{base}_values{ext}"
            
            # Load with data_only=True to get calculated values
            workbook = load_workbook(file_path, data_only=True)
            workbook.save(output_path)
            workbook.close()
            
            logger.info(f"Converted formulas to values: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to convert formulas: {str(e)}")
            raise ValueError(f"Failed to convert formulas: {str(e)}")
